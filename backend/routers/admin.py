"""Panel de administración con alcance por curso.

Roles:
- `superadmin` (jefes de la academia): acceso total. Único que gestiona cursos,
  usuarios y roles.
- `admin` (profesores): gestiona el temario/tests SOLO de los cursos que tiene
  asignados. Cada operación valida el curso implicado.
- `estudiante`: no entra aquí.
"""
import csv
import io
import os
from datetime import datetime, timedelta

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from sqlalchemy import func, cast, Integer
from sqlalchemy.orm import Session

from models import (
    get_db,
    Curso,
    Tema,
    MaterialTema,
    TestPlantilla,
    Pregunta,
    Usuario,
    RegistroFallo,
    RespuestaAlumno,
    TestIntento,
    IALlamada,
    Puntuacion,
)
from schemas import CursoIn, CursosUsuarioUpdate, TemaIn, TestPlantillaIn, PreguntaIn, RolUpdate
from routers.auth import (
    require_gestor,
    require_superadmin,
    es_superadmin,
    cursos_permitidos_ids,
    verificar_acceso_curso,
    ROLES_VALIDOS,
)
from services.preguntas import normalizar_letra

router = APIRouter(prefix="/api/admin", tags=["Administración"])

# Nombres de los usuarios de demostración que la versión antigua inyectaba en el ranking.
NOMBRES_DEMO = ["Marta V.", "Carlos M.", "Lucía Gómez", "Javier R.", "Ana Ruiz"]


# ==========================================
# HELPERS DE ALCANCE
# ==========================================
def _tema_o_404(db: Session, tema_id: int) -> Tema:
    tema = db.query(Tema).filter(Tema.id == tema_id).first()
    if not tema:
        raise HTTPException(status_code=404, detail="Tema no encontrado")
    return tema

def _temas_permitidos_ids(db: Session, usuario: Usuario):
    """IDs de temas accesibles. None = todos (superadmin)."""
    permitidos = cursos_permitidos_ids(usuario)
    if permitidos is None:
        return None
    if not permitidos:
        return []
    return [t[0] for t in db.query(Tema.id).filter(Tema.curso_id.in_(permitidos)).all()]

def _tests_permitidos_ids(db: Session, usuario: Usuario):
    """IDs de plantillas de test accesibles. None = todas (superadmin)."""
    temas = _temas_permitidos_ids(db, usuario)
    if temas is None:
        return None
    if not temas:
        return []
    return [t[0] for t in db.query(TestPlantilla.id).filter(TestPlantilla.tema_id.in_(temas)).all()]


# ==========================================
# CURSOS
# ==========================================
@router.get("/cursos")
def listar_cursos(usuario: Usuario = Depends(require_gestor), db: Session = Depends(get_db)):
    """El superadmin ve todos los cursos; el admin, solo los suyos."""
    permitidos = cursos_permitidos_ids(usuario)
    query = db.query(Curso)
    if permitidos is not None:
        query = query.filter(Curso.id.in_(permitidos or [-1]))
    return query.order_by(Curso.nombre).all()

@router.post("/cursos")
def crear_curso(datos: CursoIn, _: Usuario = Depends(require_superadmin), db: Session = Depends(get_db)):
    curso = Curso(nombre=datos.nombre.strip(), descripcion=(datos.descripcion or "").strip() or None)
    db.add(curso)
    db.commit()
    db.refresh(curso)
    return curso

@router.put("/cursos/{curso_id}")
def editar_curso(curso_id: int, datos: CursoIn, _: Usuario = Depends(require_superadmin), db: Session = Depends(get_db)):
    curso = db.query(Curso).filter(Curso.id == curso_id).first()
    if not curso:
        raise HTTPException(status_code=404, detail="Curso no encontrado")
    curso.nombre = datos.nombre.strip()
    curso.descripcion = (datos.descripcion or "").strip() or None
    db.commit()
    db.refresh(curso)
    return curso

@router.delete("/cursos/{curso_id}")
def borrar_curso(curso_id: int, _: Usuario = Depends(require_superadmin), db: Session = Depends(get_db)):
    curso = db.query(Curso).filter(Curso.id == curso_id).first()
    if not curso:
        raise HTTPException(status_code=404, detail="Curso no encontrado")
    if db.query(Tema).filter(Tema.curso_id == curso_id).first():
        raise HTTPException(status_code=400, detail="No se puede borrar: el curso todavía tiene temas.")
    db.delete(curso)
    db.commit()
    return {"mensaje": "Curso eliminado"}


# ==========================================
# TEMAS
# ==========================================
@router.get("/temas")
def listar_temas(
    curso_id: int = Query(None),
    usuario: Usuario = Depends(require_gestor),
    db: Session = Depends(get_db),
):
    permitidos = cursos_permitidos_ids(usuario)
    query = db.query(Tema)
    if permitidos is not None:
        query = query.filter(Tema.curso_id.in_(permitidos or [-1]))
    if curso_id:
        verificar_acceso_curso(usuario, curso_id)
        query = query.filter(Tema.curso_id == curso_id)
    return query.order_by(Tema.id).all()

@router.post("/temas")
def crear_tema(datos: TemaIn, usuario: Usuario = Depends(require_gestor), db: Session = Depends(get_db)):
    verificar_acceso_curso(usuario, datos.curso_id)
    tema = Tema(
        nombre=datos.nombre.strip(),
        bloque=(datos.bloque or "").strip() or None,
        curso_id=datos.curso_id,
    )
    db.add(tema)
    db.commit()
    db.refresh(tema)
    return tema

@router.put("/temas/{tema_id}")
def editar_tema(tema_id: int, datos: TemaIn, usuario: Usuario = Depends(require_gestor), db: Session = Depends(get_db)):
    tema = _tema_o_404(db, tema_id)
    verificar_acceso_curso(usuario, tema.curso_id)      # curso actual
    verificar_acceso_curso(usuario, datos.curso_id)     # curso destino
    tema.nombre = datos.nombre.strip()
    tema.bloque = (datos.bloque or "").strip() or None
    tema.curso_id = datos.curso_id
    db.commit()
    db.refresh(tema)
    return tema

@router.delete("/temas/{tema_id}")
def borrar_tema(tema_id: int, usuario: Usuario = Depends(require_gestor), db: Session = Depends(get_db)):
    tema = _tema_o_404(db, tema_id)
    verificar_acceso_curso(usuario, tema.curso_id)
    db.delete(tema)
    db.commit()
    return {"mensaje": "Tema eliminado"}


# ==========================================
# MATERIAL DEL TEMA (PDFs)
# ==========================================
# El PDF se guarda en la propia base de datos. Neon (free) son 0,5 GB en total,
# así que limitamos el tamaño por archivo.
MAX_PDF_MB = int(os.getenv("MAX_PDF_MB", "10"))

def _material_resumen(m: MaterialTema) -> dict:
    return {
        "id": m.id,
        "tema_id": m.tema_id,
        "nombre_archivo": m.nombre_archivo,
        "tamano_bytes": m.tamano_bytes,
        "fecha_subida": m.fecha_subida,
    }

@router.get("/temas/{tema_id}/materiales")
def listar_materiales(tema_id: int, usuario: Usuario = Depends(require_gestor), db: Session = Depends(get_db)):
    tema = _tema_o_404(db, tema_id)
    verificar_acceso_curso(usuario, tema.curso_id)
    return [_material_resumen(m) for m in tema.materiales]

@router.post("/temas/{tema_id}/materiales")
async def subir_material(
    tema_id: int,
    archivo: UploadFile = File(...),
    usuario: Usuario = Depends(require_gestor),
    db: Session = Depends(get_db),
):
    tema = _tema_o_404(db, tema_id)
    verificar_acceso_curso(usuario, tema.curso_id)

    nombre = (archivo.filename or "").strip()
    if not nombre.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Solo se admiten archivos PDF.")

    contenido = await archivo.read()
    if not contenido:
        raise HTTPException(status_code=400, detail="El archivo está vacío.")
    if len(contenido) > MAX_PDF_MB * 1024 * 1024:
        raise HTTPException(status_code=400, detail=f"El PDF supera el máximo permitido de {MAX_PDF_MB} MB.")

    material = MaterialTema(
        tema_id=tema_id,
        nombre_archivo=nombre,
        tipo_mime=archivo.content_type or "application/pdf",
        tamano_bytes=len(contenido),
        contenido=contenido,
    )
    db.add(material)
    db.commit()
    db.refresh(material)
    return _material_resumen(material)

@router.delete("/materiales/{material_id}")
def borrar_material(material_id: int, usuario: Usuario = Depends(require_gestor), db: Session = Depends(get_db)):
    material = db.query(MaterialTema).filter(MaterialTema.id == material_id).first()
    if not material:
        raise HTTPException(status_code=404, detail="Material no encontrado")
    verificar_acceso_curso(usuario, _tema_o_404(db, material.tema_id).curso_id)
    db.delete(material)
    db.commit()
    return {"mensaje": "Material eliminado"}


# ==========================================
# PLANTILLAS DE TEST
# ==========================================
@router.get("/tests")
def listar_tests(
    tema_id: int = Query(None),
    usuario: Usuario = Depends(require_gestor),
    db: Session = Depends(get_db),
):
    query = db.query(TestPlantilla)
    permitidos = _temas_permitidos_ids(db, usuario)
    if permitidos is not None:
        query = query.filter(TestPlantilla.tema_id.in_(permitidos or [-1]))
    if tema_id:
        verificar_acceso_curso(usuario, _tema_o_404(db, tema_id).curso_id)
        query = query.filter(TestPlantilla.tema_id == tema_id)
    return query.order_by(TestPlantilla.numero_test).all()

@router.post("/tests")
def crear_test(datos: TestPlantillaIn, usuario: Usuario = Depends(require_gestor), db: Session = Depends(get_db)):
    verificar_acceso_curso(usuario, _tema_o_404(db, datos.tema_id).curso_id)
    if db.query(TestPlantilla).filter(TestPlantilla.numero_test == datos.numero_test).first():
        raise HTTPException(status_code=400, detail="Ya existe un test con ese número")
    test = TestPlantilla(
        numero_test=datos.numero_test.strip(),
        tema_id=datos.tema_id,
        total_preguntas=datos.total_preguntas,
    )
    db.add(test)
    db.commit()
    db.refresh(test)
    return test

@router.delete("/tests/{test_id}")
def borrar_test(test_id: int, usuario: Usuario = Depends(require_gestor), db: Session = Depends(get_db)):
    test = db.query(TestPlantilla).filter(TestPlantilla.id == test_id).first()
    if not test:
        raise HTTPException(status_code=404, detail="Test no encontrado")
    verificar_acceso_curso(usuario, _tema_o_404(db, test.tema_id).curso_id)
    db.delete(test)
    db.commit()
    return {"mensaje": "Test eliminado"}


# ==========================================
# PREGUNTAS
# ==========================================
@router.get("/preguntas")
def listar_preguntas(
    test_plantilla_id: int = Query(None),
    tema_id: int = Query(None),
    usuario: Usuario = Depends(require_gestor),
    db: Session = Depends(get_db),
):
    query = db.query(Pregunta)
    permitidos = _temas_permitidos_ids(db, usuario)
    if permitidos is not None:
        query = query.filter(Pregunta.tema_id.in_(permitidos or [-1]))
    if test_plantilla_id:
        query = query.filter(Pregunta.test_plantilla_id == test_plantilla_id)
    if tema_id:
        verificar_acceso_curso(usuario, _tema_o_404(db, tema_id).curso_id)
        query = query.filter(Pregunta.tema_id == tema_id)
    return query.order_by(Pregunta.id).all()

@router.post("/preguntas")
def crear_pregunta(datos: PreguntaIn, usuario: Usuario = Depends(require_gestor), db: Session = Depends(get_db)):
    verificar_acceso_curso(usuario, _tema_o_404(db, datos.tema_id).curso_id)
    pregunta = Pregunta(
        enunciado=datos.enunciado.strip(),
        opcion_a=datos.opcion_a.strip(),
        opcion_b=datos.opcion_b.strip(),
        opcion_c=datos.opcion_c.strip(),
        opcion_d=datos.opcion_d.strip(),
        respuesta_correcta=normalizar_letra(datos.respuesta_correcta),
        explicacion=(datos.explicacion or "").strip() or None,
        tema_id=datos.tema_id,
        test_plantilla_id=datos.test_plantilla_id,
    )
    db.add(pregunta)
    db.commit()
    db.refresh(pregunta)
    return pregunta

@router.put("/preguntas/{pregunta_id}")
def editar_pregunta(pregunta_id: int, datos: PreguntaIn, usuario: Usuario = Depends(require_gestor), db: Session = Depends(get_db)):
    pregunta = db.query(Pregunta).filter(Pregunta.id == pregunta_id).first()
    if not pregunta:
        raise HTTPException(status_code=404, detail="Pregunta no encontrada")
    verificar_acceso_curso(usuario, _tema_o_404(db, pregunta.tema_id).curso_id)
    verificar_acceso_curso(usuario, _tema_o_404(db, datos.tema_id).curso_id)
    pregunta.enunciado = datos.enunciado.strip()
    pregunta.opcion_a = datos.opcion_a.strip()
    pregunta.opcion_b = datos.opcion_b.strip()
    pregunta.opcion_c = datos.opcion_c.strip()
    pregunta.opcion_d = datos.opcion_d.strip()
    pregunta.respuesta_correcta = normalizar_letra(datos.respuesta_correcta)
    pregunta.explicacion = (datos.explicacion or "").strip() or None
    pregunta.tema_id = datos.tema_id
    pregunta.test_plantilla_id = datos.test_plantilla_id
    db.commit()
    db.refresh(pregunta)
    return pregunta

@router.delete("/preguntas/{pregunta_id}")
def borrar_pregunta(pregunta_id: int, usuario: Usuario = Depends(require_gestor), db: Session = Depends(get_db)):
    pregunta = db.query(Pregunta).filter(Pregunta.id == pregunta_id).first()
    if not pregunta:
        raise HTTPException(status_code=404, detail="Pregunta no encontrada")
    verificar_acceso_curso(usuario, _tema_o_404(db, pregunta.tema_id).curso_id)
    db.delete(pregunta)
    db.commit()
    return {"mensaje": "Pregunta eliminada"}


# ==========================================
# IMPORTACIÓN MASIVA (CSV / XLSX)
# ==========================================
def _filas_desde_csv(contenido: bytes):
    texto = contenido.decode("utf-8-sig")
    return [dict(fila) for fila in csv.DictReader(io.StringIO(texto))]

def _filas_desde_xlsx(contenido: bytes):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="Falta la librería openpyxl para leer .xlsx. Usa un CSV o instálala.")
    wb = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    filas = list(wb.active.iter_rows(values_only=True))
    if not filas:
        return []
    cabeceras = [str(c).strip() if c is not None else "" for c in filas[0]]
    return [
        {cabeceras[i]: fila[i] for i in range(len(cabeceras)) if i < len(fila)}
        for fila in filas[1:]
    ]

@router.post("/preguntas/importar")
async def importar_preguntas(
    archivo: UploadFile = File(...),
    usuario: Usuario = Depends(require_gestor),
    db: Session = Depends(get_db),
):
    """Importa preguntas desde CSV/XLSX con columnas: tema_id, numero_test,
    enunciado, opcion_a..d, respuesta_correcta, explicacion.
    Solo se aceptan filas de temas a los que tengas acceso."""
    contenido = await archivo.read()
    nombre = (archivo.filename or "").lower()

    if nombre.endswith(".csv"):
        filas = _filas_desde_csv(contenido)
    elif nombre.endswith(".xlsx"):
        filas = _filas_desde_xlsx(contenido)
    else:
        raise HTTPException(status_code=400, detail="Formato no soportado. Sube un .csv o .xlsx")

    creadas = 0
    errores = []
    for i, fila in enumerate(filas, start=2):  # fila 1 = cabecera
        try:
            tema_id = int(fila["tema_id"])
            tema = db.query(Tema).filter(Tema.id == tema_id).first()
            if not tema:
                raise ValueError(f"el tema {tema_id} no existe")
            verificar_acceso_curso(usuario, tema.curso_id)

            numero_test = str(int(float(fila["numero_test"]))).zfill(3)
            plantilla = db.query(TestPlantilla).filter(
                TestPlantilla.tema_id == tema_id,
                TestPlantilla.numero_test == numero_test,
            ).first()
            if not plantilla:
                plantilla = TestPlantilla(tema_id=tema_id, numero_test=numero_test, total_preguntas=10)
                db.add(plantilla)
                db.commit()
                db.refresh(plantilla)

            db.add(Pregunta(
                tema_id=tema_id,
                test_plantilla_id=plantilla.id,
                enunciado=str(fila["enunciado"]).strip(),
                opcion_a=str(fila["opcion_a"]).strip(),
                opcion_b=str(fila["opcion_b"]).strip(),
                opcion_c=str(fila["opcion_c"]).strip(),
                opcion_d=str(fila["opcion_d"]).strip(),
                respuesta_correcta=normalizar_letra(fila["respuesta_correcta"]),
                explicacion=(str(fila.get("explicacion") or "").strip() or "Consulta el temario para más detalle."),
            ))
            creadas += 1
        except HTTPException as e:
            db.rollback()
            errores.append(f"Fila {i}: {e.detail}")
        except Exception as e:
            db.rollback()
            errores.append(f"Fila {i}: {e}")

    db.commit()
    return {"creadas": creadas, "errores": errores}


# ==========================================
# MÉTRICAS (limitadas a los cursos accesibles)
# ==========================================
@router.get("/metricas")
def obtener_metricas(usuario: Usuario = Depends(require_gestor), db: Session = Depends(get_db)):
    hace_7d = datetime.utcnow() - timedelta(days=7)
    hoy = datetime.utcnow().date()

    temas_ok = _temas_permitidos_ids(db, usuario)
    tests_ok = _tests_permitidos_ids(db, usuario)
    cursos_ok = cursos_permitidos_ids(usuario)

    # --- Usuarios ---
    if es_superadmin(usuario):
        usuarios_total = db.query(func.count(Usuario.id)).scalar() or 0
        alumnos_total = db.query(func.count(Usuario.id)).filter(Usuario.rol == "estudiante").scalar() or 0
    else:
        # Solo los estudiantes matriculados en los cursos del profesor
        alumnos_total = (
            db.query(func.count(func.distinct(Usuario.id)))
            .join(Usuario.cursos)
            .filter(Curso.id.in_(cursos_ok or [-1]), Usuario.rol == "estudiante")
            .scalar() or 0
        )
        usuarios_total = alumnos_total

    # --- Alumnos activos (7d) ---
    q_activos = db.query(func.count(func.distinct(TestIntento.alumno_id))).filter(TestIntento.fecha_intento >= hace_7d)
    if tests_ok is not None:
        q_activos = q_activos.filter(TestIntento.test_plantilla_id.in_(tests_ok or [-1]))
    alumnos_activos_7d = q_activos.scalar() or 0

    # --- Progreso medio por tema ---
    q_prog = (
        db.query(
            Tema.id,
            Tema.nombre,
            func.count(RespuestaAlumno.id),
            func.sum(cast(RespuestaAlumno.es_correcta, Integer)),
        )
        .join(Pregunta, Pregunta.tema_id == Tema.id)
        .join(RespuestaAlumno, RespuestaAlumno.pregunta_id == Pregunta.id)
    )
    if temas_ok is not None:
        q_prog = q_prog.filter(Tema.id.in_(temas_ok or [-1]))
    progreso_por_tema = []
    for tema_id, nombre, total, correctas in q_prog.group_by(Tema.id, Tema.nombre).all():
        total = total or 0
        correctas = correctas or 0
        progreso_por_tema.append({
            "tema_id": tema_id,
            "nombre": nombre,
            "respuestas": total,
            "porcentaje": round((correctas / total) * 100) if total else 0,
        })

    # --- Preguntas más falladas ---
    q_fallos = db.query(RegistroFallo.pregunta_id, func.count(RegistroFallo.id).label("veces"))
    if temas_ok is not None:
        preguntas_ok = [p[0] for p in db.query(Pregunta.id).filter(Pregunta.tema_id.in_(temas_ok or [-1])).all()]
        q_fallos = q_fallos.filter(RegistroFallo.pregunta_id.in_(preguntas_ok or [-1]))
    preguntas_mas_falladas = []
    for pregunta_id, veces in q_fallos.group_by(RegistroFallo.pregunta_id).order_by(func.count(RegistroFallo.id).desc()).limit(10).all():
        p = db.query(Pregunta).filter(Pregunta.id == pregunta_id).first()
        preguntas_mas_falladas.append({
            "pregunta_id": pregunta_id,
            "enunciado": p.enunciado if p else "(pregunta eliminada)",
            "veces_fallada": veces,
        })

    # --- Uso de IA (solo lo ve el superadmin: es un coste global) ---
    if es_superadmin(usuario):
        ia = {
            "total": db.query(func.count(IALlamada.id)).scalar() or 0,
            "ultimos_7d": db.query(func.count(IALlamada.id)).filter(IALlamada.fecha >= hace_7d).scalar() or 0,
            "hoy": db.query(func.count(IALlamada.id)).filter(func.date(IALlamada.fecha) == hoy).scalar() or 0,
            "tokens_totales": db.query(func.coalesce(func.sum(IALlamada.tokens_totales), 0)).scalar() or 0,
        }
    else:
        ia = None

    return {
        "usuarios_total": usuarios_total,
        "alumnos_total": alumnos_total,
        "alumnos_activos_7d": alumnos_activos_7d,
        "progreso_por_tema": progreso_por_tema,
        "preguntas_mas_falladas": preguntas_mas_falladas,
        "ia": ia,
    }


# ==========================================
# RANKING: limpieza
# ==========================================
@router.delete("/ranking/demo")
def limpiar_ranking_demo(_: Usuario = Depends(require_gestor), db: Session = Depends(get_db)):
    borradas = db.query(Puntuacion).filter(Puntuacion.alumno_nombre.in_(NOMBRES_DEMO)).delete(synchronize_session=False)
    db.commit()
    return {"mensaje": f"Eliminadas {borradas} puntuaciones de demostración."}

@router.delete("/ranking/reset")
def reset_ranking(_: Usuario = Depends(require_superadmin), db: Session = Depends(get_db)):
    borradas = db.query(Puntuacion).delete(synchronize_session=False)
    db.commit()
    return {"mensaje": f"Ranking reiniciado. Eliminadas {borradas} puntuaciones."}


# ==========================================
# USUARIOS, ROLES Y CURSOS (solo superadmin)
# ==========================================
@router.get("/usuarios")
def listar_usuarios(_: Usuario = Depends(require_superadmin), db: Session = Depends(get_db)):
    return [
        {
            "id": u.id,
            "nombre": u.nombre,
            "email": u.email,
            "rol": u.rol,
            "cursos": [{"id": c.id, "nombre": c.nombre} for c in u.cursos],
        }
        for u in db.query(Usuario).order_by(Usuario.id).all()
    ]

@router.put("/usuarios/{usuario_id}/rol")
def cambiar_rol(
    usuario_id: int,
    datos: RolUpdate,
    admin: Usuario = Depends(require_superadmin),
    db: Session = Depends(get_db),
):
    if datos.rol not in ROLES_VALIDOS:
        raise HTTPException(status_code=400, detail=f"Rol no válido. Usa uno de: {', '.join(ROLES_VALIDOS)}")
    usuario = db.query(Usuario).filter(Usuario.id == usuario_id).first()
    if not usuario:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    if usuario.id == admin.id and datos.rol != "superadmin":
        raise HTTPException(status_code=400, detail="No puedes quitarte a ti mismo el rol de superadministrador.")
    usuario.rol = datos.rol
    db.commit()
    return {"id": usuario.id, "nombre": usuario.nombre, "email": usuario.email, "rol": usuario.rol}

@router.put("/usuarios/{usuario_id}/cursos")
def asignar_cursos(
    usuario_id: int,
    datos: CursosUsuarioUpdate,
    _: Usuario = Depends(require_superadmin),
    db: Session = Depends(get_db),
):
    """Asigna los cursos de un usuario: los que gestiona (admin) o en los que
    está matriculado (estudiante)."""
    usuario = db.query(Usuario).filter(Usuario.id == usuario_id).first()
    if not usuario:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    cursos = db.query(Curso).filter(Curso.id.in_(datos.curso_ids or [-1])).all()
    if len(cursos) != len(set(datos.curso_ids or [])):
        raise HTTPException(status_code=400, detail="Algún curso indicado no existe.")
    usuario.cursos = cursos
    db.commit()
    return {
        "id": usuario.id,
        "nombre": usuario.nombre,
        "rol": usuario.rol,
        "cursos": [{"id": c.id, "nombre": c.nombre} for c in usuario.cursos],
    }
